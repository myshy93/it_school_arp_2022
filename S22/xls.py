import pathlib
from openpyxl import Workbook, load_workbook
# https://openpyxl.readthedocs.io/en/stable/

root = pathlib.Path(__file__).parent
files_path = root.joinpath("files")

cantitati = {}

# read xlsx files
wb = load_workbook(files_path.joinpath("inventory.xlsx")) # deschide fisierul pentru citire
# print(type(wb))
# print(wb)
# print(wb.sheetnames)
sheet1 = wb['Sheet']


rows = list(sheet1.iter_rows()) # extragem toate liniile
rows = rows[1:] # eliminam headerul

for i in rows:
    if int(i[2].value) < 30:
        cantitati[i[0].value] = i[2].value

print(cantitati)
wb.close()


# write xlsx files
filename = "list_cumparaturi.xlsx"

# init workbook
workbook = Workbook()
sheet = workbook.active

# header - how to highlight
sheet["A1"] = "Produs"
sheet["B1"] = "U.M."
sheet["C1"] = "Cantitate"

sheet["A2"] = "Vopsea"
sheet["B2"] = "BUC"
sheet["C2"] = 10

list_sheet = workbook.create_sheet('list')

for i in cantitati.keys():
    list_sheet.append([i])

workbook.save(filename=files_path.joinpath(filename))